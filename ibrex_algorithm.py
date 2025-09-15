# -*- coding: utf-8 -*-
"""
Created on Wed Jul  6 18:02:24 2022

@author: gowrishankar.p
"""

import os,openpyxl
from openpyxl import load_workbook

data = ['SUBTRACT']

path = 'D:\gowrishankar.p\Python Script\Inp_script\COBOL-Phrase-Mapping.xlsx'
wb=openpyxl.load_workbook(path)
sheets = wb.sheetnames

def mapping():
    conf  = 'D:\gowrishankar.p\Python Script\Inp_script\mapping.txt'
    # global configd
    configd = {}
    proc_ok = False 
    try:
      config = open (conf,'r')
      proc_ok = True 
    except:
        print ("mapping file mapping.txt is missing !! ")

    if proc_ok:
        for line in config:
            keys,val = line.strip().split('<=>',1)
            configd.update(dict.fromkeys(keys.split('<=>'),val))
        config.close()
    return (configd)

def GetCellData(rown, coln):
    cell_obj = sheet_obj.cell(row=rown, column=coln)
    return cell_obj.value

def conversion(var):
    split = var.split()
    line =[]
    sub_var = ''
    from_Var = ''
    giv_var = ''
    giving = False
   ## To find the varaibales between the strings ##
    
    start = "SUBTRACT"
    end = "FROM"
    end1 =["GIVING",".","END"]
    
    # To find the value between Sub **** From
    sub_var = var[var.find(start)+len(start):var.rfind(end)]
    var_split = sub_var.split()
    # print("first ",var[var.find(start)+len(start):var.rfind(end)])
    
    
    # Find the final brk point for varaiables
    for e in end1:
        if e in var:
            secondend = e
            break
    
    # To find the value between from **** 'giving'/'end'/'.' parameters
    from_var = var[var.find(end)+len(end):var.rfind(secondend)]
    var_split1 = from_var.split()
    # print("second ",var[var.find(end)+len(end):var.rfind(secondend)])
    
    start1 = 'GIVING'
    if var.find('GIVING') > 0:
        giv_var = var[var.find(start1)+len(start1):len(var)]
        # print("third ",var[var.find(start1)+len(start1):len(var)])
        giving = True
    
    if len(var_split) == 1 and giving == False:
        data = sub_var + "will be subtracted from " + from_var + ' and the results will be stored in' \
            + from_var
        line.append(data)
    elif len(var_split) > 1 and giving == False:
        data = "Numbers/Values" + sub_var + 'will be summed up  and the results subtracted from' \
            + from_var + 'and stored in same data' + from_var
        line.append(data)
        # print(data)
    
    if len(var_split) == 1 and giving == True:
        data = sub_var + "will be subtracted from " + from_var + ' and the results will be stored in' \
            + giv_var
        line.append(data)
    elif len(var_split) > 1 and giving == True:
        data = "Numbers/Values" + sub_var + 'will be summed up  and the results subtracted from' \
            + from_var + 'and stored in same data' + giv_var
        line.append(data)
        
    
    return(line)
        
def conversion2(var):
    split = var.split()
    line =[]
    sub_var = ''
    from_Var = ''
    giv_var = ''
    giving = False
   ## To find the varaibales between the strings ##
    
    start = "SUBTRACT"
    end = "FROM"
    end1 =["GIVING",".","END"]
    
    # To find the value between Sub **** From
    sub_var = var[var.find(start)+len(start):var.rfind(end)]
    var_split = sub_var.split()
    # print("first ",var[var.find(start)+len(start):var.rfind(end)])
    
    
    # Find the final brk point for varaiables
    for e in end1:
        if e in var:
            secondend = e
            break
    
    # To find the value between from **** 'giving'/'end'/'.' parameters
    from_var = var[var.find(end)+len(end):var.rfind(secondend)]
    var_split1 = from_var.split()
    # print("second ",var[var.find(end)+len(end):var.rfind(secondend)])
    
    start1 = 'GIVING'
    if var.find('GIVING') > 0:
        giv_var = var[var.find(start1)+len(start1):len(var)]
        # print("third ",var[var.find(start1)+len(start1):len(var)])
        giving = True
    
    for each in split:
        if each in config:
            # print(config[each])
            line.append(' ' +config[each])
        else:
            # print(each)
            line.append(' '+each)
    if giving == False:    
        if var.find('END') >= 0 and  var.find('.') >= 0:
            line.append(from_var)
        else:
            line.append('store the result in'+from_var)

    
    return(line)
        
def read_xls():
    global sheet_obj
    for sheet in sheets:
        sheet_obj = wb[sheet]
        if sheet == "COBOL-Phrase-Mapping":
            for row1 in range(2,sheet_obj.max_row+1):
                    var = GetCellData(row1,2)
                    print(var)
                    
                    ## Method 1
                    ret = conversion(var)
                    print('#1 >>',' '.join(ret))
                    
                    ## Method 2
                    ret = conversion2(var)
                    print('#2 >>',' '.join(ret),"\n")
                    
                    if row1 == 6:
                        break

                
def main():
    global config
    config = mapping()
    # print(config['SUBTRACT'])
    read_xls()
if __name__ == "__main__":
    main()