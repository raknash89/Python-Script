# -*- coding: utf-8 -*-
"""
Created on Wed Jun 29 18:45:09 2022

@author: gowrishankar.p
"""
import openpyxl,os,sys

path = "D:\gowrishankar.p\DB2 to SQL Migrator tool\Load_Script_Requirement_test.xlsx"
# path2 = "D:\gowrishankar.p\DB2 to SQL Migrator tool\dataset\G176281.SPUNCH.D220622.TBL1.DAT"
path2 = "D:\gowrishankar.p\DB2 to SQL Migrator tool\dataset\DATA.UNLD.LRECL.txt"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
sheets = wb_obj.sheetnames
linestrip=[]
with open(path2, "r") as script:
    lines = script.readlines()
    # lines = lines.replace('\n','')
    for a in lines:
        linestrip.append(a.rstrip('\n'))
    
print("------------")

for sheet in sheets:
        sheet_obj = wb_obj[sheet]
        if sheet == "LRECL":
            i = 1
            for each in linestrip:
                # each = each.strip('\n')
                # print(i)
                print(each)
                var = sheet_obj.cell(row=sheet_obj.max_row+i,column=1)
                print(var.row)
                var.value = each
                print(var.value)
                wb_obj.save(path)
            # for row1 in range(2,sheet_obj.max_row+1):
            #     # var = sheet_obj.cell(row = 2, column = 1)
            #     var = sheet_obj.cell(row=row1,column=1)
            #     print(var.value)
                # var_split = var.split(",")
                # list_of_lrecl_dsn.append(var_split[0])
                # list_of_lrecl_dsn_format.append(var_split[1].strip())
                # list_of_lrecl_dsn_type.append(var_split[2].strip())
                # list_of_lrecl_reclen.append(var_split[3])
                
wb_obj.close()

# scriptPath = "D:\gowrishankar.p\DB2 to SQL Migrator tool\dataset\UNLOAD_LRECL.txt"


# cell_obj = sheet_obj.cell(row = 1, column = 1)
# print(cell_obj.value)